"""
Export functionality for salary reports
Supports Excel, CSV, and formatted text exports
"""
import os
import csv
from typing import Dict, Any, List, Optional
from datetime import datetime
import pandas as pd

try:
    # Optional: Try to import openpyxl for Excel export
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from models import BonusInfo


class ReportExporter:
    """Class for exporting salary calculation reports in various formats"""
    
    def __init__(self):
        self.excel_available = EXCEL_AVAILABLE
    
    def export_to_csv(self, filepath: str, detailed_df: pd.DataFrame, 
                     grouped_df: pd.DataFrame, salary_data: Dict[str, Any]) -> bool:
        """
        Export report to CSV format
        
        Args:
            filepath: Output file path
            detailed_df: Detailed flight data
            grouped_df: Grouped daily data
            salary_data: Salary calculation results
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create a summary of the export
            export_data = []
            
            # Add salary summary
            export_data.append(['=== SALARY SUMMARY ==='])
            export_data.append(['Gross Total', f"{salary_data.get('gross_total', 0):.2f} €"])
            export_data.append(['Net Estimated', f"{salary_data.get('net_estimated', 0):.2f} €"])
            export_data.append(['Operational Sectors', f"{salary_data.get('operational_sectors_earnings', 0):.2f} €"])
            export_data.append(['Positioning Flights', f"{salary_data.get('positioning_earnings', 0):.2f} €"])
            export_data.append([''])
            
            # Add daily summary
            export_data.append(['=== DAILY SUMMARY ==='])
            export_data.append(['Date', 'Activity', 'Flights', 'Sectors', 'Earnings'])
            
            for _, row in grouped_df.iterrows():
                export_data.append([
                    row['Data'].strftime('%Y-%m-%d'),
                    row['Attività'],
                    row['Volo'],
                    f"{row['Settori']:.2f}",
                    f"{row['Guadagno (€)']:.2f}"
                ])
            
            export_data.append([''])
            
            # Add detailed flight data
            export_data.append(['=== DETAILED FLIGHTS ==='])
            export_data.append(['Date', 'Flight', 'Origin', 'Destination', 'Distance', 'Sectors', 'Earnings', 'Type'])
            
            for _, row in detailed_df.iterrows():
                export_data.append([
                    row['Data'].strftime('%Y-%m-%d'),
                    row['Volo'],
                    row['Partenza'],
                    row['Arrivo'],
                    f"{row['Distanza']:.0f}" if row['Distanza'] > 0 else '---',
                    f"{row['Settori']:.2f}",
                    f"{row['Guadagno (€)']:.2f}",
                    'Positioning' if row['IsPositioning'] else 'Flight'
                ])
            
            # Write to CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(export_data)
            
            return True
            
        except Exception:
            return False
    
    def export_to_excel(self, filepath: str, detailed_df: pd.DataFrame, 
                       grouped_df: pd.DataFrame, salary_data: Dict[str, Any],
                       ido_bonuses: List[BonusInfo], extra_diaria_days: set,
                       profile_data: Dict[str, Any]) -> bool:
        """
        Export report to Excel format with formatting
        
        Args:
            filepath: Output file path
            detailed_df: Detailed flight data
            grouped_df: Grouped daily data  
            salary_data: Salary calculation results
            ido_bonuses: IDO bonus information
            extra_diaria_days: Days with extra diaria
            profile_data: Pilot profile information
        
        Returns:
            True if successful, False otherwise
        """
        if not self.excel_available:
            return False
        
        try:
            workbook = openpyxl.Workbook()
            
            # Create summary sheet
            self._create_summary_sheet(workbook, salary_data, profile_data)
            
            # Create daily schedule sheet
            self._create_schedule_sheet(workbook, grouped_df, ido_bonuses, extra_diaria_days)
            
            # Create detailed flights sheet
            self._create_details_sheet(workbook, detailed_df)
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Save workbook
            workbook.save(filepath)
            return True
            
        except Exception:
            return False
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, 
                             salary_data: Dict[str, Any], profile_data: Dict[str, Any]):
        """Create summary sheet in Excel workbook"""
        ws = workbook.active
        ws.title = "Salary Summary"
        
        # Styles
        header_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        currency_font = Font(bold=True, color="2F5496")
        
        # Header
        ws['A1'] = "PILOT SALARY CALCULATION SUMMARY"
        ws['A1'].font = header_font
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Profile information
        ws[f'A{row}'] = "Profile Information"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = "Position:"
        ws[f'B{row}'] = profile_data.get('position', '')
        row += 1
        
        ws[f'A{row}'] = "Contract:"
        ws[f'B{row}'] = profile_data.get('contract_type', '')
        row += 1
        
        ws[f'A{row}'] = "Home Base:"
        ws[f'B{row}'] = profile_data.get('home_base', '')
        row += 2
        
        # Salary breakdown
        ws[f'A{row}'] = "Salary Components"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        salary_items = [
            ("Gross Total Salary", salary_data.get('gross_total', 0)),
            ("Social Contributions", -salary_data.get('social_contributions', 0)),
            ("Taxable Income", salary_data.get('taxable_income', 0)),
            ("Estimated Tax", -salary_data.get('estimated_tax', 0)),
            ("Net Estimated Salary", salary_data.get('net_estimated', 0)),
        ]
        
        for label, value in salary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"{value:.2f} €"
            if "Net Estimated" in label:
                ws[f'B{row}'].font = currency_font
            row += 1
        
        row += 1
        
        # Earnings breakdown
        ws[f'A{row}'] = "Earnings Breakdown"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        earnings_items = [
            ("Operational Sectors", salary_data.get('operational_sectors_earnings', 0)),
            ("Positioning Flights", salary_data.get('positioning_earnings', 0)),
            ("FRV Bonus", salary_data.get('frv_bonus', 0)),
            ("SNC Compensation", salary_data.get('snc_compensation', 0)),
            ("Vacation Pay", salary_data.get('vacation_compensation', 0)),
        ]
        
        for label, value in earnings_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"{value:.2f} €"
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def _create_schedule_sheet(self, workbook: openpyxl.Workbook, grouped_df: pd.DataFrame,
                              ido_bonuses: List[BonusInfo], extra_diaria_days: set):
        """Create daily schedule sheet"""
        ws = workbook.create_sheet("Daily Schedule")
        
        # Headers
        headers = ['Date', 'Activity', 'Flights', 'Sectors', 'Earnings', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row_idx, (_, row) in enumerate(grouped_df.iterrows(), 2):
            date_str = row['Data'].strftime('%Y-%m-%d')
            
            ws.cell(row=row_idx, column=1, value=date_str)
            ws.cell(row=row_idx, column=2, value=row['Attività'])
            ws.cell(row=row_idx, column=3, value=row['Volo'])
            ws.cell(row=row_idx, column=4, value=f"{row['Settori']:.2f}")
            ws.cell(row=row_idx, column=5, value=f"{row['Guadagno (€)']:.2f} €")
            
            # Add notes for bonuses
            notes = []
            if date_str in extra_diaria_days:
                notes.append("Extra Diaria")
            
            for bonus in ido_bonuses:
                if bonus.date == date_str:
                    notes.append(f"IDO Bonus {bonus.symbol}")
            
            if notes:
                ws.cell(row=row_idx, column=6, value=", ".join(notes))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)
    
    def _create_details_sheet(self, workbook: openpyxl.Workbook, detailed_df: pd.DataFrame):
        """Create detailed flights sheet"""
        ws = workbook.create_sheet("Flight Details")
        
        # Add DataFrame to worksheet
        for r in dataframe_to_rows(detailed_df, index=False, header=True):
            ws.append(r)
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 25)
    
    def export_to_text(self, filepath: str, grouped_df: pd.DataFrame, 
                      salary_data: Dict[str, Any], profile_data: Dict[str, Any]) -> bool:
        """
        Export report to formatted text file
        
        Args:
            filepath: Output file path
            grouped_df: Grouped daily data
            salary_data: Salary calculation results
            profile_data: Pilot profile information
        
        Returns:
            True if successful, False otherwise
        """
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                # Header
                f.write("=" * 60 + "\n")
                f.write("PILOT SALARY CALCULATION REPORT\n")
                f.write("=" * 60 + "\n\n")
                
                # Generation info
                f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Position: {profile_data.get('position', 'N/A')}\n")
                f.write(f"Contract: {profile_data.get('contract_type', 'N/A')}\n")
                f.write(f"Home Base: {profile_data.get('home_base', 'N/A')}\n\n")
                
                # Salary summary
                f.write("SALARY SUMMARY\n")
                f.write("-" * 60 + "\n")
                f.write(f"{'Gross Total Salary:':<35} {salary_data.get('gross_total', 0):>15,.2f} €\n")
                f.write(f"{'Social Contributions:':<35} {-salary_data.get('social_contributions', 0):>15,.2f} €\n")
                f.write(f"{'Taxable Income:':<35} {salary_data.get('taxable_income', 0):>15,.2f} €\n")
                f.write(f"{'Estimated Tax:':<35} {-salary_data.get('estimated_tax', 0):>15,.2f} €\n")
                f.write("-" * 60 + "\n")
                f.write(f"{'NET ESTIMATED SALARY:':<35} {salary_data.get('net_estimated', 0):>15,.2f} €\n\n")
                
                # Earnings breakdown
                f.write("EARNINGS BREAKDOWN\n")
                f.write("-" * 60 + "\n")
                f.write(f"{'Operational Sectors:':<35} {salary_data.get('operational_sectors_earnings', 0):>15,.2f} €\n")
                f.write(f"{'Positioning Flights:':<35} {salary_data.get('positioning_earnings', 0):>15,.2f} €\n")
                f.write(f"{'FRV Bonus:':<35} {salary_data.get('frv_bonus', 0):>15,.2f} €\n")
                f.write(f"{'SNC Compensation:':<35} {salary_data.get('snc_compensation', 0):>15,.2f} €\n")
                f.write(f"{'Vacation Pay:':<35} {salary_data.get('vacation_compensation', 0):>15,.2f} €\n\n")
                
                # Daily schedule
                f.write("DAILY SCHEDULE\n")
                f.write("-" * 80 + "\n")
                f.write(f"{'Date':<12} {'Activity':<25} {'Flights':<8} {'Sectors':<8} {'Earnings':<12}\n")
                f.write("-" * 80 + "\n")
                
                for _, row in grouped_df.iterrows():
                    date_str = row['Data'].strftime('%Y-%m-%d')
                    activity = row['Attività'][:24]  # Truncate if too long
                    flights = str(row['Volo']) if row['Volo'] != '---' else '---'
                    sectors = f"{row['Settori']:.2f}"
                    earnings = f"{row['Guadagno (€)']:.2f} €"
                    
                    f.write(f"{date_str:<12} {activity:<25} {flights:<8} {sectors:<8} {earnings:<12}\n")
                
                f.write("-" * 80 + "\n")
                f.write(f"Report generated by Pilot Salary Calculator v2.0\n")
            
            return True
            
        except Exception:
            return False